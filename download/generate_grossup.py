import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Kalkulator"

# ===== STYLES =====
header_fill = PatternFill(start_color="1B3F66", end_color="1B3F66", fill_type="solid")
subheader_fill = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
header_font = Font(name='Times New Roman', color="FFFFFF", bold=True, size=11)
subheader_font = Font(name='Times New Roman', color="FFFFFF", bold=True, size=10)
title_font = Font(name='Times New Roman', size=18, bold=True, color="000000")
subtitle_font = Font(name='Times New Roman', size=11, color="666666", italic=True)
section_font = Font(name='Times New Roman', size=13, bold=True, color="1B3F66")
label_font = Font(name='Times New Roman', size=11, color="000000")
input_font = Font(name='Times New Roman', size=11, color="0000FF")
formula_font = Font(name='Times New Roman', size=11, color="000000")
bold_formula_font = Font(name='Times New Roman', size=11, bold=True, color="000000")
result_font = Font(name='Times New Roman', size=12, bold=True, color="1B3F66")
note_font = Font(name='Times New Roman', size=10, color="666666", italic=True)
check_font = Font(name='Times New Roman', size=11, bold=True, color="008000")
input_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
result_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
alt_fill = PatternFill(start_color="E9E9E9", end_color="E9E9E9", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

thin_side = Side(style='thin', color='C0C0C0')
medium_side = Side(style='medium', color='1B3F66')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
bottom_thick = Border(bottom=medium_side)
no_border = Border()

left_al = Alignment(horizontal='left', vertical='center')
center_al = Alignment(horizontal='center', vertical='center')
right_al = Alignment(horizontal='right', vertical='center')
wrap_al = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ===== COLUMN WIDTHS =====
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 28
ws.column_dimensions['E'].width = 28
ws.column_dimensions['F'].width = 24

# ===== TITLE =====
ws.row_dimensions[2].height = 35
ws['B2'] = "TEMPLATE GROSS-UP PPh 21 TAHUNAN"
ws['B2'].font = title_font
ws['B2'].alignment = left_al

ws['B3'] = "PPh dibebankan oleh Pemberi Kerja (Tax Gross-Up Method)"
ws['B3'].font = subtitle_font
ws['B3'].alignment = left_al

# ===== SECTION 1: DATA INPUT =====
r = 5
ws.merge_cells(f'B{r}:C{r}')
ws[f'B{r}'] = "DATA INPUT"
ws[f'B{r}'].font = header_font
ws[f'B{r}'].fill = header_fill
ws[f'B{r}'].alignment = left_al
ws[f'C{r}'].fill = header_fill
ws.row_dimensions[r].height = 26

input_data = [
    (6, "Nama Karyawan", "", None, None, ""),
    (7, "Status PTKP (TK/K)", "TK", "TK,K", None, "TK = Tidak Kawin, K = Kawin"),
    (8, "Jumlah Tanggungan (0-3)", 0, "0,1,2,3", None, "Maksimal 3 tanggungan"),
    (9, "Penghasilan Bersih (Net) / Tahun (Rp)", 120000000, None, '#,##0', "Take-home pay setelah PPh"),
    (10, "Tarif Biaya Jabatan (%)", 0.05, None, '0%', "Default 5%, maks Rp 6.000.000/thn"),
    (11, "Iuran Pensiun Karyawan / Tahun (Rp)", 0, None, '#,##0', "Opsional, mengurangi PKP"),
    (12, "Iuran BPJS Kesehatan Karyawan / Tahun (Rp)", 0, None, '#,##0', "Opsional, mengurangi PKP"),
]

for row, label, default, dv_list, fmt, note in input_data:
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = wrap_al
    ws.cell(row=row, column=2).border = thin_border

    cell = ws.cell(row=row, column=3, value=default)
    cell.font = input_font
    cell.alignment = right_al
    cell.border = thin_border
    cell.fill = input_fill
    if fmt:
        cell.number_format = fmt

    if dv_list:
        dv = DataValidation(type="list", formula1=f'"{dv_list}"', allow_blank=False)
        dv.error = "Pilih dari daftar yang tersedia"
        dv.errorTitle = "Input Tidak Valid"
        dv.prompt = f"Pilih: {dv_list}"
        dv.promptTitle = "Pilihan"
        ws.add_data_validation(dv)
        dv.add(cell)

    if note:
        ws.cell(row=row, column=4, value=note).font = note_font
        ws.cell(row=row, column=4).alignment = left_al

# ===== SECTION 2: KALKULASI PTKP =====
r = 14
ws.merge_cells(f'B{r}:C{r}')
ws[f'B{r}'] = "PERHITUNGAN PTKP"
ws[f'B{r}'].font = header_font
ws[f'B{r}'].fill = header_fill
ws[f'B{r}'].alignment = left_al
ws[f'C{r}'].fill = header_fill
ws.row_dimensions[r].height = 26

ws.cell(row=15, column=2, value="Dasar PTKP").font = label_font
ws.cell(row=15, column=2).border = thin_border
ws.cell(row=15, column=2).alignment = left_al
c = ws.cell(row=15, column=3)
c.value = '=IF(C7="TK",54000000,58500000)'
c.font = formula_font
c.border = thin_border
c.alignment = right_al
c.number_format = '#,##0'

ws.cell(row=16, column=2, value="Tambahan Tanggungan").font = label_font
ws.cell(row=16, column=2).border = thin_border
ws.cell(row=16, column=2).alignment = left_al
c = ws.cell(row=16, column=3)
c.value = '=C8*4500000'
c.font = formula_font
c.border = thin_border
c.alignment = right_al
c.number_format = '#,##0'

ws.cell(row=17, column=2, value="PTKP Total / Tahun").font = Font(name='Times New Roman', size=11, bold=True, color="000000")
ws.cell(row=17, column=2).border = thin_border
ws.cell(row=17, column=2).fill = result_fill
ws.cell(row=17, column=2).alignment = left_al
c = ws.cell(row=17, column=3)
c.value = '=C15+C16'
c.font = bold_formula_font
c.border = thin_border
c.fill = result_fill
c.alignment = right_al
c.number_format = '#,##0'

# ===== SECTION 3: ITERASI GROSS-UP =====
r = 19
ws.merge_cells(f'B{r}:F{r}')
ws[f'B{r}'] = "PROSES ITERASI GROSS-UP"
ws[f'B{r}'].font = header_font
ws[f'B{r}'].fill = header_fill
ws[f'B{r}'].alignment = left_al
for col in range(3, 7):
    ws.cell(row=r, column=col).fill = header_fill
ws.row_dimensions[r].height = 26

iter_headers = ["Iterasi", "Penghasilan Bruto (Rp)", "Biaya Jabatan (Rp)", "PKP (Rp)", "PPh 21 (Rp)"]
for col_idx, hdr in enumerate(iter_headers, 2):
    c = ws.cell(row=20, column=col_idx, value=hdr)
    c.font = subheader_font
    c.fill = subheader_fill
    c.alignment = center_al
    c.border = thin_border

def pph_formula(pkp_ref):
    return (
        f'=IF({pkp_ref}<=0,0,'
        f'MIN({pkp_ref},60000000)*0.05+'
        f'MAX(MIN({pkp_ref}-60000000,190000000),0)*0.15+'
        f'MAX(MIN({pkp_ref}-250000000,250000000),0)*0.25+'
        f'MAX(MIN({pkp_ref}-500000000,4500000000),0)*0.3+'
        f'MAX({pkp_ref}-5000000000,0)*0.35)'
    )

num_iterations = 12
for i in range(num_iterations):
    row = 21 + i
    fill = alt_fill if i % 2 == 1 else white_fill

    ws.cell(row=row, column=2, value=f"Iterasi {i}").font = Font(name='Times New Roman', size=10, color="333333")
    ws.cell(row=row, column=2).alignment = center_al
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).fill = fill

    # Gross: Iter 0 = Net, others = Net + PPh_prev
    if i == 0:
        gross = '=C9'
    else:
        gross = f'=$C$9+F{row-1}'

    c = ws.cell(row=row, column=3, value=gross)
    c.font = formula_font
    c.alignment = right_al
    c.border = thin_border
    c.fill = fill
    c.number_format = '#,##0'

    # Biaya Jabatan
    c = ws.cell(row=row, column=4, value=f'=MIN(C{row}*$C$10,6000000)')
    c.font = formula_font
    c.alignment = right_al
    c.border = thin_border
    c.fill = fill
    c.number_format = '#,##0'

    # PKP
    c = ws.cell(row=row, column=5, value=f'=MAX(C{row}-D{row}-$C$17-$C$11-$C$12,0)')
    c.font = formula_font
    c.alignment = right_al
    c.border = thin_border
    c.fill = fill
    c.number_format = '#,##0'

    # PPh
    c = ws.cell(row=row, column=6, value=pph_formula(f'E{row}'))
    c.font = formula_font
    c.alignment = right_al
    c.border = thin_border
    c.fill = fill
    c.number_format = '#,##0'

last_iter_row = 21 + num_iterations - 1  # 28

# ===== SECTION 4: HASIL AKHIR =====
r = last_iter_row + 2  # 30
ws.merge_cells(f'B{r}:D{r}')
ws[f'B{r}'] = "HASIL AKHIR"
ws[f'B{r}'].font = header_font
ws[f'B{r}'].fill = header_fill
ws[f'B{r}'].alignment = left_al
for col in range(3, 5):
    ws.cell(row=r, column=col).fill = header_fill
ws.row_dimensions[r].height = 26

# Rincian Perhitungan
r_detail = r + 1  # 31
ws[f'B{r_detail}'] = "Rincian Perhitungan"
ws[f'B{r_detail}'].font = section_font
ws[f'B{r_detail}'].alignment = left_al

detail_rows = [
    (r_detail+1, "Penghasilan Bruto (Gross) / Tahun", f'=C{last_iter_row}', '#,##0'),
    (r_detail+2, "PTKP", '=C17', '#,##0'),
    (r_detail+3, "Biaya Jabatan", f'=D{last_iter_row}', '#,##0'),
    (r_detail+4, "Iuran Pensiun Karyawan", '=C11', '#,##0'),
    (r_detail+5, "Iuran BPJS Kesehatan Karyawan", '=C12', '#,##0'),
    (r_detail+6, "PKP (Penghasilan Kena Pajak)", f'=E{last_iter_row}', '#,##0'),
]

for row, label, formula, fmt in detail_rows:
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = left_al
    c = ws.cell(row=row, column=3, value=formula)
    c.font = formula_font
    c.border = thin_border
    c.alignment = right_al
    c.number_format = fmt

# Ringkasan
r_summary = r_detail + 8  # 40
ws[f'B{r_summary}'] = "Ringkasan"
ws[f'B{r_summary}'].font = section_font
ws[f'B{r_summary}'].alignment = left_al

summary_rows = [
    (r_summary+1, "PPh 21 Terutang / Tahun", f'=F{last_iter_row}', '#,##0', white_fill),
    (r_summary+2, "Effective Tax Rate", f'=IF(C{last_iter_row}>0,F{last_iter_row}/C{last_iter_row},0)', '0.00%', white_fill),
    (r_summary+3, "Penghasilan Bersih (Net) / Tahun", f'=C{last_iter_row}-F{last_iter_row}', '#,##0', result_fill),
    (r_summary+4, "", None, None, white_fill),  # separator
    (r_summary+5, "Penghasilan Bruto / Bulan", f'=C{last_iter_row}/12', '#,##0', white_fill),
    (r_summary+6, "PPh 21 / Bulan", f'=F{last_iter_row}/12', '#,##0', white_fill),
    (r_summary+7, "Penghasilan Bersih (Net) / Bulan", f'=(C{last_iter_row}-F{last_iter_row})/12', '#,##0', result_fill),
    (r_summary+8, "", None, None, white_fill),  # separator
    (r_summary+9, "Total Biaya Majikan / Tahun", f'=C{last_iter_row}+F{last_iter_row}', '#,##0', total_fill),
    (r_summary+10, "Total Biaya Majikan / Bulan", f'=(C{last_iter_row}+F{last_iter_row})/12', '#,##0', total_fill),
]

for row, label, formula, fmt, fill in summary_rows:
    if not label:
        continue
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = left_al
    if fill != white_fill:
        ws.cell(row=row, column=2).fill = fill
    c = ws.cell(row=row, column=3, value=formula)
    c.font = bold_formula_font if fill != white_fill else formula_font
    c.border = thin_border
    c.alignment = right_al
    c.number_format = fmt
    if fill != white_fill:
        c.fill = fill

# Verification
r_verif = r_summary + 12
ws.cell(row=r_verif, column=2, value="Verifikasi (Net Hasil = Net Input)").font = label_font
ws.cell(row=r_verif, column=2).border = thin_border
ws.cell(row=r_verif, column=2).alignment = left_al
c = ws.cell(row=r_verif, column=3)
c.value = f'=IF(ABS(C{r_summary+3}-C9)<=1,"✓ COCOK","✗ TIDAK COCOK - Selisih: "&TEXT(ABS(C{r_summary+3}-C9),"#,##0"))'
c.font = check_font
c.border = thin_border
c.alignment = center_al

# ===== SHEET 2: REFERENSI =====
ws2 = wb.create_sheet("Referensi")

ws2.column_dimensions['A'].width = 3
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['E'].width = 35
ws2.column_dimensions['F'].width = 18

# Title
ws2['B2'] = "TABEL REFERENSI PPh 21"
ws2['B2'].font = title_font
ws2['B2'].alignment = left_al
ws2.row_dimensions[2].height = 35

# PTKP Table
ws2.merge_cells('B4:C4')
ws2['B4'] = "TABEL PTKP (TAHUNAN)"
ws2['B4'].font = header_font
ws2['B4'].fill = header_fill
ws2['B4'].alignment = left_al
ws2['C4'].fill = header_fill

ptkp_data = [
    ("TK/0", 54000000),
    ("TK/1", 58500000),
    ("TK/2", 63000000),
    ("TK/3", 67500000),
    ("K/0", 58500000),
    ("K/1", 63000000),
    ("K/2", 67500000),
    ("K/3", 72000000),
]

ws2['B5'] = "Status"
ws2['C5'] = "PTKP (Rp/Tahun)"
for col in ['B', 'C']:
    ws2[f'{col}5'].font = subheader_font
    ws2[f'{col}5'].fill = subheader_fill
    ws2[f'{col}5'].alignment = center_al
    ws2[f'{col}5'].border = thin_border

for i, (status, amount) in enumerate(ptkp_data):
    row = 6 + i
    fill = alt_fill if i % 2 == 1 else white_fill
    c = ws2.cell(row=row, column=2, value=status)
    c.font = label_font
    c.alignment = center_al
    c.border = thin_border
    c.fill = fill
    c = ws2.cell(row=row, column=3, value=amount)
    c.font = formula_font
    c.alignment = right_al
    c.border = thin_border
    c.number_format = '#,##0'
    c.fill = fill

ws2['B15'] = "Keterangan:"
ws2['B15'].font = Font(name='Times New Roman', size=10, bold=True, color="333333")
ws2['B16'] = "TK = Tidak Kawin  |  K = Kawin"
ws2['B16'].font = note_font
ws2['B17'] = "Dasar PTKP TK: Rp 54.000.000 | K: Rp 58.500.000"
ws2['B17'].font = note_font
ws2['B18'] = "Tambahan per tanggungan: Rp 4.500.000"
ws2['B18'].font = note_font

# Tarif PPh Table
ws2.merge_cells('E4:F4')
ws2['E4'] = "TARIF PROGRESIF PPh 21"
ws2['E4'].font = header_font
ws2['E4'].fill = header_fill
ws2['E4'].alignment = left_al
ws2['F4'].fill = header_fill

ws2['E5'] = "Lapisan PKP (Rp/Tahun)"
ws2['F5'] = "Tarif"
for col in ['E', 'F']:
    ws2[f'{col}5'].font = subheader_font
    ws2[f'{col}5'].fill = subheader_fill
    ws2[f'{col}5'].alignment = center_al
    ws2[f'{col}5'].border = thin_border

tarif_data = [
    ("0 - 60.000.000", "5%"),
    ("60.000.000 - 250.000.000", "15%"),
    ("250.000.000 - 500.000.000", "25%"),
    ("500.000.000 - 5.000.000.000", "30%"),
    ("> 5.000.000.000", "35%"),
]

for i, (lapisan, tarif) in enumerate(tarif_data):
    row = 6 + i
    fill = alt_fill if i % 2 == 1 else white_fill
    c = ws2.cell(row=row, column=5, value=lapisan)
    c.font = label_font
    c.alignment = left_al
    c.border = thin_border
    c.fill = fill
    c = ws2.cell(row=row, column=6, value=tarif)
    c.font = Font(name='Times New Roman', size=11, bold=True, color="000000")
    c.alignment = center_al
    c.border = thin_border
    c.fill = fill

# Notes section
r_note = 13
ws2[f'E{r_note}'] = "Catatan Penting:"
ws2[f'E{r_note}'].font = Font(name='Times New Roman', size=10, bold=True, color="333333")

notes = [
    "1. Biaya Jabatan: 5% dari penghasilan bruto, maks Rp 500.000/bulan (Rp 6.000.000/tahun)",
    "2. PTKP dan tarif berdasarkan UU HPP No. 7 Tahun 2021 (berlaku sejak 2022)",
    "3. Iuran pensiun/BPJS yang ditanggung karyawan mengurangi PKP",
    "4. Gross-up: PPh dibebankan oleh pemberi kerja, bukan dipotong dari gaji karyawan",
    "5. Formula iterasi: Gross = Net + PPh, hingga konvergen (biasanya 4-6 iterasi)",
]
for i, note in enumerate(notes):
    ws2[f'E{r_note+1+i}'] = note
    ws2[f'E{r_note+1+i}'].font = note_font

# ===== PRINT SETTINGS =====
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = 'portrait'

# ===== FREEZE PANES =====
ws.freeze_panes = 'B5'

# ===== SAVE =====
output_path = '/home/z/my-project/download/Template_Gross_Up_PPh21.xlsx'
wb.save(output_path)
print(f"File saved: {output_path}")
